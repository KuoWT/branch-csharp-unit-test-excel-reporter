#!/usr/bin/env python3
"""Generate the department-style C# NUnit DLL test report as a modern .xlsx."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from zipfile import ZipFile
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
BLUE = PatternFill("solid", fgColor="D9EAF7")
GRAY = PatternFill("solid", fgColor="E7E6E6")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PACKAGE_RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def load_json(path: str) -> dict:
    if not path:
        return {}
    p = Path(path)
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))


def style_range(ws, cell_range: str, fill=None, bold=False):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill
            if bold:
                cell.font = Font(bold=True)


def result_fill(value):
    text = str(value or "").lower()
    if text in {"pass", "yes", "可 push"}:
        return GREEN
    if text in {"fail", "no", "missing", "inconclusive", "不可 push"}:
        return RED
    return YELLOW


def validate_xlsx(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    if not wb.sheetnames:
        raise ValueError(f"{path} has no worksheets.")

    with ZipFile(path) as archive:
        bad_member = archive.testzip()
        if bad_member:
            raise ValueError(f"{path} contains a corrupt zip member: {bad_member}")

        for rels_path in ("_rels/.rels", "xl/_rels/workbook.xml.rels"):
            root = ET.fromstring(archive.read(rels_path))
            if root.tag != f"{{{PACKAGE_RELS_NS}}}Relationships":
                raise ValueError(f"{rels_path} uses an invalid relationships namespace.")


def collect_scenarios(diff: dict, tests: list[dict]) -> list[dict]:
    if tests:
        return tests
    rows = []
    for item in diff.get("files", []):
        methods = item.get("methods") or [{"class": "", "function": "(unknown)"}]
        for method in methods:
            rows.append(
                {
                    "file": item.get("file", ""),
                    "class": method.get("class", ""),
                    "function": method.get("function", ""),
                    "change_type": item.get("change_type", "Unknown"),
                    "positive_case": "",
                    "positive_result": "Missing",
                    "negative_case": "",
                    "negative_result": "Missing",
                    "nunit_test": "",
                    "notes": "Test case requires implementation.",
                }
            )
    return rows


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--diff-json", required=True)
    parser.add_argument("--build-json", default="")
    parser.add_argument("--tests-json", default="")
    parser.add_argument("--output", required=True)
    parser.add_argument("--site", default="")
    parser.add_argument("--request-no", default="")
    parser.add_argument("--requirement", default="")
    parser.add_argument("--user-tester", default="")
    parser.add_argument("--it-tester", default="")
    parser.add_argument("--customer", default="")
    parser.add_argument("--package", default="")
    args = parser.parse_args()

    diff = load_json(args.diff_json)
    build = load_json(args.build_json)
    tests = load_json(args.tests_json).get("tests", []) if args.tests_json else []
    scenarios = collect_scenarios(diff, tests)

    build_pass = build.get("overall") == "Pass"
    positive_ok = all(str(s.get("positive_result", "")).lower() == "pass" for s in scenarios) and bool(scenarios)
    negative_ok = all(str(s.get("negative_result", "")).lower() == "pass" for s in scenarios) and bool(scenarios)
    push_allowed = build_pass and positive_ok and negative_ok

    wb = Workbook()
    ws = wb.active
    ws.title = "IT測試報告"
    detail = wb.create_sheet("測試明細")
    result = wb.create_sheet("執行結果")

    ws.merge_cells("A1:K1")
    ws["A1"] = "IT 測試報告"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = BLUE

    fields = [
        ("站點", args.site),
        ("IT影響評估", f"Base Tag {diff.get('base_tag', '')} 至 Branch {diff.get('branch', '')}，異動 C# function 共 {len(scenarios)} 筆。"),
        ("申請單號", args.request_no),
        ("需求內容", args.requirement or f"依 {diff.get('diff_range', '')} diff 產生 NUnit 正反相測試報告。"),
        ("User 測試人員", args.user_tester),
        ("限定客戶", args.customer),
        ("限定PACKAGE", args.package),
        ("IT 測試人員", args.it_tester),
    ]
    row = 3
    for label, value in fields:
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
        row += 1
    style_range(ws, f"A3:K{row-1}", fill=None)
    for r in range(3, row):
        ws.cell(r, 1).fill = GRAY
        ws.cell(r, 1).font = Font(bold=True)

    row += 1
    ws.cell(row, 1, "情境")
    ws.cell(row, 1).font = Font(bold=True, size=13)
    headers = ["編號", "類型", "檔案", "類別", "Function", "測試情境", "NUnit Test", "預期結果", "實際結果", "結果", "備註"]
    row += 1
    for col, header in enumerate(headers, 1):
        ws.cell(row, col, header)
    style_range(ws, f"A{row}:K{row}", fill=BLUE, bold=True)
    start = row + 1
    index = 1
    for s in scenarios:
        for kind, case_key, result_key in [
            ("Positive", "positive_case", "positive_result"),
            ("Negative", "negative_case", "negative_result"),
        ]:
            ws.append(
                [
                    index,
                    kind,
                    s.get("file", ""),
                    s.get("class", ""),
                    s.get("function", ""),
                    s.get(case_key, ""),
                    s.get("nunit_test", ""),
                    "Pass expected behavior" if kind == "Positive" else "Reject invalid behavior",
                    s.get(result_key, ""),
                    s.get(result_key, "Missing"),
                    s.get("notes", ""),
                ]
            )
            index += 1
    end = ws.max_row
    if end >= start:
        style_range(ws, f"A{start}:K{end}")
        for r in range(start, end + 1):
            ws.cell(r, 10).fill = result_fill(ws.cell(r, 10).value)

    detail_headers = [
        "Branch",
        "Base Tag",
        "Commit",
        "File",
        "Class",
        "Function",
        "Change Type",
        "Positive Case",
        "Positive Result",
        "Negative Case",
        "Negative Result",
        "Notes",
    ]
    detail.append(detail_headers)
    for s in scenarios:
        detail.append(
            [
                diff.get("branch", ""),
                diff.get("base_tag", ""),
                diff.get("commit", ""),
                s.get("file", ""),
                s.get("class", ""),
                s.get("function", ""),
                s.get("change_type", ""),
                s.get("positive_case", ""),
                s.get("positive_result", "Missing"),
                s.get("negative_case", ""),
                s.get("negative_result", "Missing"),
                s.get("notes", ""),
            ]
        )
    style_range(detail, f"A1:L{max(detail.max_row, 1)}")
    style_range(detail, "A1:L1", fill=BLUE, bold=True)
    detail.auto_filter.ref = detail.dimensions
    detail.freeze_panes = "A2"
    for r in range(2, detail.max_row + 1):
        detail.cell(r, 9).fill = result_fill(detail.cell(r, 9).value)
        detail.cell(r, 11).fill = result_fill(detail.cell(r, 11).value)

    result_rows = [
        ("Branch", diff.get("branch", ""), "指定本地分支"),
        ("Base Tag", diff.get("base_tag", ""), "最新 tag"),
        ("Commit", diff.get("commit", ""), "分支 HEAD"),
        ("Restore", build.get("restore", {}).get("result", "Unknown"), ""),
        ("Build", build.get("build", {}).get("result", "Unknown"), ""),
        ("NUnit Test", build.get("test", {}).get("result", "Unknown"), ""),
        ("正向測試覆蓋", "Pass" if positive_ok else "Fail", "每個 function 必須有正向測試且 pass"),
        ("反向測試覆蓋", "Pass" if negative_ok else "Fail", "每個 function 必須有反向測試且 pass"),
        ("Excel Report", "Pass", args.output),
        ("Push Allowed", "Yes" if push_allowed else "No", "全部 gate pass 才允許 push"),
    ]
    result.append(["項目", "結果", "備註"])
    for item in result_rows:
        result.append(list(item))
    style_range(result, f"A1:C{result.max_row}")
    style_range(result, "A1:C1", fill=BLUE, bold=True)
    for r in range(2, result.max_row + 1):
        result.cell(r, 2).fill = result_fill(result.cell(r, 2).value)
    result.auto_filter.ref = result.dimensions
    result.freeze_panes = "A2"

    widths = {
        "IT測試報告": [8, 12, 24, 18, 22, 34, 28, 24, 24, 12, 28],
        "測試明細": [14, 14, 18, 28, 18, 22, 14, 30, 16, 30, 16, 30],
        "執行結果": [20, 20, 50],
    }
    for sheet in wb.worksheets:
        for idx, width in enumerate(widths.get(sheet.title, []), 1):
            sheet.column_dimensions[get_column_letter(idx)].width = width
        for cells in sheet.iter_rows():
            for cell in cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    validate_xlsx(output)
    print(f"Saved {output}")
    print("PASS: 可 push" if push_allowed else "FAIL: 不可 push")
    return 0 if push_allowed else 2


if __name__ == "__main__":
    raise SystemExit(main())
